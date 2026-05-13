"""
Maestro de personas. Carga desde CSV/XLSX y upsert masivo.

Acepta cabeceras flexibles:
  employee_no | person id | id           → employee_no  (PK)
  nombre      | name                     → nombre
  departamento| department | depto       → departamento
  activo      | active                   → activo (opcional, default TRUE)
"""
from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass
from typing import Iterable

from . import db

log = logging.getLogger("persona")

# ---------- normalización de cabeceras ----------

_HEADER_ALIASES = {
    "employee_no": {"employee_no", "personid", "person_id", "id", "legajo"},
    "nombre":      {"nombre", "name", "fullname", "full_name"},
    "departamento":{"departamento", "department", "depto", "dept"},
    "activo":      {"activo", "active", "enabled"},
}


def _norm(s: str) -> str:
    return re.sub(r"[\s_\-]+", "", (s or "").lstrip("\ufeff").strip().lower())


def _map_headers(headers: list[str]) -> dict[str, int]:
    """devuelve {campo_canónico: índice_columna}"""
    out: dict[str, int] = {}
    for i, h in enumerate(headers):
        nh = _norm(h)
        for canon, aliases in _HEADER_ALIASES.items():
            if nh in {_norm(a) for a in aliases}:
                out.setdefault(canon, i)
                break
    return out


# ---------- parseo ----------

@dataclass
class PersonaRow:
    employee_no: str
    nombre: str
    departamento: str | None
    activo: bool


def _coerce_activo(v) -> bool:
    if v is None:
        return True
    s = str(v).strip().lower()
    if s in {"", "1", "true", "t", "si", "sí", "y", "yes", "activo"}:
        return True
    return s not in {"0", "false", "f", "no", "n", "inactivo"}


def parse_csv(text: str) -> list[PersonaRow]:
    # sniff delimitador (coma o punto y coma)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rdr = csv.reader(io.StringIO(text), dialect)
    rows = list(rdr)
    if not rows:
        return []
    return _parse_rows(rows)


def parse_xlsx(blob: bytes) -> list[PersonaRow]:
    # import perezoso: si no se usa la carga XLSX, no se necesita el dep
    from openpyxl import load_workbook  # type: ignore
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb.active
    rows = [[("" if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]
    return _parse_rows(rows)


def _norm_emp_no(raw: str) -> str:
    """
    Normaliza el ID al formato que el reloj guarda en asistencia.evento: padded a 8 dígitos.
    El CSV exportado trae apóstrofe Excel-escape ("'00000001"). El reloj devuelve
    employeeNoString también padded ("00000001"). IDs numéricos cortos se completan a 8.
    Si no es numérico, se devuelve tal cual (sin el apóstrofe).
    """
    s = str(raw).strip().lstrip("'").strip()
    if s.isdigit():
        return s.zfill(8)
    return s


def _parse_rows(rows: list[list[str]]) -> list[PersonaRow]:
    if not rows:
        return []
    header = rows[0]
    header_idx = _map_headers(header)
    missing = {"employee_no", "nombre"} - header_idx.keys()
    if missing:
        raise ValueError(f"faltan columnas obligatorias: {sorted(missing)}. "
                         f"cabeceras vistas: {header}")

    # Detección de fila rota por ';' dentro del Name:
    # El export de Hikvision tiene el patrón [Person ID; Name; Department; ] (4 cols con trailing).
    # Si Name contenía ';', la fila aparece como 4 valores [ID, "Apellido", " Nombre", "0 min"]
    # → la 4ta col deja de estar vacía. Reconstruyo: name = col[1]+" "+col[2], dept = None.
    name_col = header_idx["nombre"]
    dept_col = header_idx.get("departamento")
    n_header = len(header)

    out: list[PersonaRow] = []
    seen: set[str] = set()
    for r in rows[1:]:
        if not any((c or "").strip() for c in r):
            continue
        try:
            emp = _norm_emp_no(r[header_idx["employee_no"]])
            nom = str(r[name_col]).strip()
        except IndexError:
            continue
        depto = str(r[dept_col]).strip() if dept_col is not None and dept_col < len(r) else ""

        # caso "fila rota": cuando la última col del header (trailing) trae datos
        # significa que un ';' se metió en el nombre y desplazó todo
        if n_header >= 4 and len(r) >= n_header and r[n_header - 1].strip():
            nom = (str(r[name_col]).strip() + " " + str(r[name_col + 1]).strip()).strip()
            depto = ""  # se perdió en el shift; queda NULL

        if not emp or not nom:
            continue
        if emp in seen:  # dedup intra-archivo, último gana
            out = [p for p in out if p.employee_no != emp]
        seen.add(emp)

        activo = True
        if "activo" in header_idx and header_idx["activo"] < len(r):
            activo = _coerce_activo(r[header_idx["activo"]])
        out.append(PersonaRow(emp, nom, (depto or None), activo))
    return out


# ---------- persistencia ----------

def upsert(personas: Iterable[PersonaRow]) -> int:
    personas = list(personas)
    if not personas:
        return 0
    rows = [(p.employee_no, p.nombre, p.departamento, p.activo) for p in personas]
    with db.get_conn() as c, c.cursor() as cur:
        cur.executemany("""
            INSERT INTO asistencia.persona (employee_no, nombre, departamento, activo, updated_at)
            VALUES (%s, %s, %s, %s, NOW())
            ON CONFLICT (employee_no) DO UPDATE
              SET nombre       = EXCLUDED.nombre,
                  departamento = EXCLUDED.departamento,
                  activo       = EXCLUDED.activo,
                  updated_at   = NOW()
        """, rows)
    log.info(f"persona upsert: {len(rows)} filas")
    return len(rows)


def count() -> int:
    with db.get_conn() as c, c.cursor() as cur:
        cur.execute("SELECT COUNT(*) AS n FROM asistencia.persona")
        return cur.fetchone()["n"]
